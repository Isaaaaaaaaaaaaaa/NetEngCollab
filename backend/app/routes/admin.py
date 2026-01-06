from datetime import datetime, timedelta

from sqlalchemy import or_

from flask import Blueprint, jsonify, request
from flask_jwt_extended import get_jwt_identity, jwt_required

from ..extensions import db
from ..models import (
    CooperationProject,
    CooperationRequest,
    CooperationStatus,
    Message,
    Resource,
    ReviewStatus,
    Role,
    StudentProfile,
    TeacherPost,
    User,
)
from ..rbac import require_roles
from ..utils import json_loads, now_utc


bp = Blueprint("admin", __name__)


@bp.post("/users/<int:user_id>/set-password")
@require_roles(["admin"])
def set_user_password(user_id: int):
    from ..utils import hash_password

    data = request.get_json(force=True)
    password = data.get("password") or ""
    if not password:
        return jsonify({"message": "参数不完整"}), 400
    u = User.query.get(user_id)
    if not u:
        return jsonify({"message": "用户不存在"}), 404
    u.password_hash = hash_password(password)
    u.must_change_password = password == "123456"
    db.session.commit()
    return jsonify({"ok": True})


@bp.post("/cooperations/<int:req_id>/release")
@require_roles(["admin"])
def release_cooperation(req_id: int):
    req = CooperationRequest.query.get(req_id)
    if not req:
        return jsonify({"message": "不存在"}), 404
    req.teacher_status = CooperationStatus.rejected.value
    req.student_status = CooperationStatus.rejected.value
    req.final_status = CooperationStatus.rejected.value
    req.updated_at = now_utc()
    db.session.commit()
    return jsonify({"ok": True})


@bp.post("/cooperations/<int:req_id>/reset")
@require_roles(["admin"])
def reset_cooperation(req_id: int):
    req = CooperationRequest.query.get(req_id)
    if not req:
        return jsonify({"message": "不存在"}), 404
    proj = CooperationProject.query.filter_by(request_id=req.id).first()
    if proj:
        db.session.delete(proj)
    db.session.delete(req)
    db.session.commit()
    return jsonify({"ok": True})


@bp.get("/stats")
@require_roles(["admin"])
def stats():
    return jsonify(
        {
            "users": User.query.count(),
            "teacher_posts": TeacherPost.query.count(),
            "resources": Resource.query.count(),
        }
    )


@bp.get("/analytics")
@require_roles(["admin"])
def analytics():
    now = now_utc()

    # 日维度：最近 14 天发布量 / 沟通量
    days_span = 14
    start_day = (now - timedelta(days=days_span - 1)).date()

    posts_recent = TeacherPost.query.filter(TeacherPost.created_at >= datetime.combine(start_day, datetime.min.time())).all()
    messages_recent = Message.query.filter(Message.created_at >= datetime.combine(start_day, datetime.min.time())).all()

    def count_by_date(rows, attr="created_at"):
        m = {}
        for r in rows:
            d = getattr(r, attr).date().isoformat()
            m[d] = m.get(d, 0) + 1
        result = []
        for i in range(days_span):
            d = (start_day + timedelta(days=i)).isoformat()
            result.append({"date": d, "count": m.get(d, 0)})
        return result

    posts_daily = count_by_date(posts_recent)
    messages_daily = count_by_date(messages_recent)

    # 月维度：最近 6 个月发布量 / 沟通量
    def month_key(dt: datetime) -> str:
        return dt.strftime("%Y-%m")

    months_span = 6
    # 取最近 6 个月的第一天
    first_of_this_month = datetime(now.year, now.month, 1)
    start_month_dt = first_of_this_month - timedelta(days=30 * (months_span - 1))

    posts_recent_month = TeacherPost.query.filter(TeacherPost.created_at >= start_month_dt).all()
    messages_recent_month = Message.query.filter(Message.created_at >= start_month_dt).all()

    def count_by_month(rows):
        m = {}
        for r in rows:
            k = month_key(r.created_at)
            m[k] = m.get(k, 0) + 1
        result = []
        cur = first_of_this_month
        for i in range(months_span):
            k = cur.strftime("%Y-%m")
            result.append({"month": k, "count": m.get(k, 0)})
            # 上一月
            if cur.month == 1:
                cur = datetime(cur.year - 1, 12, 1)
            else:
                cur = datetime(cur.year, cur.month - 1, 1)
        result.reverse()
        return result

    posts_monthly = count_by_month(posts_recent_month)
    messages_monthly = count_by_month(messages_recent_month)

    # 热门方向：统计教师项目的 tags / tech_stack 出现频次
    tag_counter = {}
    all_posts = TeacherPost.query.all()
    for p in all_posts:
        tags = json_loads(p.tags_json, []) or []
        techs = json_loads(p.tech_stack_json, []) or []
        for t in list(tags) + list(techs):
            if not t:
                continue
            name = str(t)
            tag_counter[name] = tag_counter.get(name, 0) + 1
    hot_topics = [
        {"name": name, "count": count}
        for name, count in sorted(tag_counter.items(), key=lambda x: x[1], reverse=True)[:10]
    ]

    # 竞赛参与趋势：近 6 个月内，post_type=competition 且 final_status=confirmed 的合作数
    coop_q = CooperationRequest.query.filter_by(final_status=CooperationStatus.confirmed.value).all()
    comp_counter = {}
    for r in coop_q:
        post = TeacherPost.query.get(r.post_id) if r.post_id else None
        if not post or post.post_type != "competition":
            continue
        k = month_key(r.created_at)
        comp_counter[k] = comp_counter.get(k, 0) + 1
    competition_trend = [
        {"month": m["month"], "count": comp_counter.get(m["month"], 0)} for m in posts_monthly
    ]

    return jsonify(
        {
            "posts_daily": posts_daily,
            "messages_daily": messages_daily,
            "posts_monthly": posts_monthly,
            "messages_monthly": messages_monthly,
            "hot_topics": hot_topics,
            "competition_trend": competition_trend,
        }
    )


@bp.get("/pending-users")
@require_roles(["admin"])
def pending_users():
    users = User.query.filter_by(is_active=False).order_by(User.created_at.desc()).all()
    return jsonify(
        {
            "items": [
                {
                    "id": u.id,
                    "username": u.username,
                    "display_name": u.display_name,
                    "role": u.role,
                    "created_at": u.created_at.isoformat(),
                }
                for u in users
            ]
        }
    )


@bp.get("/pending-teacher-posts")
@require_roles(["admin"])
def pending_teacher_posts():
    posts = (
        TeacherPost.query.filter_by(review_status=ReviewStatus.pending.value)
        .order_by(TeacherPost.created_at.desc())
        .all()
    )
    items = []
    for p in posts:
        teacher = User.query.get(p.teacher_user_id)
        items.append(
            {
                "id": p.id,
                "post_type": p.post_type,
                "title": p.title,
                "content": p.content,
                "created_at": p.created_at.isoformat(),
                "teacher": {
                    "id": teacher.id,
                    "display_name": teacher.display_name,
                }
                if teacher
                else None,
            }
        )
    return jsonify({"items": items})


@bp.get("/pending-resources")
@require_roles(["admin"])
def pending_resources():
    resources = (
        Resource.query.filter_by(review_status=ReviewStatus.pending.value)
        .order_by(Resource.created_at.desc())
        .all()
    )
    items = []
    for r in resources:
        uploader = User.query.get(r.uploader_user_id)
        items.append(
            {
                "id": r.id,
                "title": r.title,
                "resource_type": r.resource_type,
                "created_at": r.created_at.isoformat(),
                "uploader": {
                    "id": uploader.id,
                    "display_name": uploader.display_name,
                }
                if uploader
                else None,
            }
        )
    return jsonify({"items": items})


@bp.get("/users")
@require_roles(["admin"])
def list_users():
    role = (request.args.get("role") or "").strip()
    keyword = (request.args.get("keyword") or "").strip()
    page = max(int(request.args.get("page", 1)), 1)
    page_size = int(request.args.get("page_size", 20))
    if page_size <= 0 or page_size > 100:
        page_size = 20
    q = User.query
    if role in {"student", "teacher", "admin"}:
        q = q.filter_by(role=role)
    if keyword:
        q = q.filter(or_(User.username.contains(keyword), User.display_name.contains(keyword)))
    total = q.count()
    users = (
        q.order_by(User.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return jsonify(
        {
            "items": [
                {
                    "id": u.id,
                    "username": u.username,
                    "display_name": u.display_name,
                    "role": u.role,
                    "is_active": u.is_active,
                    "created_at": u.created_at.isoformat() if u.created_at else None,
                    "email": u.email,
                    "phone": u.phone,
                }
                for u in users
            ],
            "total": total,
            "page": page,
            "page_size": page_size,
        }
    )


@bp.get("/users/<int:user_id>")
@require_roles(["admin"])
def get_user_detail(user_id: int):
    u = User.query.get(user_id)
    if not u:
        return jsonify({"message": "用户不存在"}), 404

    base = {
        "id": u.id,
        "username": u.username,
        "display_name": u.display_name,
        "role": u.role,
        "is_active": u.is_active,
        "created_at": u.created_at.isoformat() if u.created_at else None,
        "email": u.email,
        "phone": u.phone,
    }

    detail = {"user": base}

    if u.role == Role.student.value:
        p = StudentProfile.query.filter_by(user_id=u.id).first()
        if p:
            detail["student_profile"] = {
                "major": p.major,
                "grade": p.grade,
                "class_name": p.class_name,
                "weekly_hours": p.weekly_hours,
                "prefer_local": p.prefer_local,
                "accept_cross": p.accept_cross,
                "visibility": p.visibility,
                "skills": json_loads(p.skills_json, []),
                "interests": json_loads(p.interests_json, []),
                "project_links": json_loads(p.project_links_json, []),
                "experiences": json_loads(p.experiences_json or "[]", []),
            }
        else:
            detail["student_profile"] = None

        detail["stats"] = {
            "confirmed_projects": CooperationRequest.query.filter_by(
                student_user_id=u.id,
                final_status=CooperationStatus.confirmed.value,
            ).count(),
        }

    if u.role == Role.teacher.value:
        detail["stats"] = {
            "published_posts": TeacherPost.query.filter_by(teacher_user_id=u.id).count(),
            "confirmed_projects": CooperationRequest.query.filter_by(
                teacher_user_id=u.id,
                final_status=CooperationStatus.confirmed.value,
            ).count(),
        }

    return jsonify(detail)


@bp.post("/users/batch-create")
@require_roles(["admin"])
def batch_create_users():
    from ..utils import hash_password, now_utc

    data = request.get_json(force=True)
    entries = data.get("users") or []
    created = []
    for e in entries:
        username = (e.get("username") or "").strip()
        password = e.get("password") or "123456"
        role = e.get("role") or "student"
        display_name = (e.get("display_name") or username).strip()
        if not username or role not in {"student", "teacher", "admin"}:
            continue
        if User.query.filter_by(username=username).first():
            continue
        u = User(
            username=username,
            password_hash=hash_password(password),
            role=role,
            display_name=display_name,
            is_active=True,
            created_at=now_utc(),
            must_change_password=password == "123456",
        )
        db.session.add(u)
        created.append({"username": username, "role": role})
    db.session.commit()
    return jsonify({"created": created})


@bp.get("/cooperations")
@require_roles(["admin"])
def cooperation_overview():
    reqs = CooperationRequest.query.order_by(CooperationRequest.created_at.desc()).limit(500).all()
    items = []
    for r in reqs:
        teacher = User.query.get(r.teacher_user_id)
        student = User.query.get(r.student_user_id)
        post = TeacherPost.query.get(r.post_id) if r.post_id else None
        items.append(
            {
                "id": r.id,
                "teacher": {"id": teacher.id, "display_name": teacher.display_name} if teacher else None,
                "student": {"id": student.id, "display_name": student.display_name} if student else None,
                "post": {"id": post.id, "title": post.title} if post else None,
                "teacher_status": r.teacher_status,
                "student_status": r.student_status,
                "final_status": r.final_status,
                "created_at": r.created_at.isoformat(),
            }
        )
    summary = {
        "total": len(reqs),
        "confirmed": len([r for r in reqs if r.final_status == CooperationStatus.confirmed.value]),
        "rejected": len([r for r in reqs if r.final_status == CooperationStatus.rejected.value]),
        "pending": len([r for r in reqs if r.final_status == CooperationStatus.pending.value]),
    }
    return jsonify({"items": items, "summary": summary})


@bp.get("/projects")
@require_roles(["admin"])
def list_projects():
    page = max(int(request.args.get("page", 1)), 1)
    page_size = int(request.args.get("page_size", 20))
    if page_size <= 0 or page_size > 100:
        page_size = 20

    post_type = (request.args.get("post_type") or "").strip()
    keyword = (request.args.get("keyword") or "").strip()

    q = TeacherPost.query
    if post_type in {"project", "innovation", "competition"}:
        q = q.filter_by(post_type=post_type)
    if keyword:
        q = q.filter(TeacherPost.title.contains(keyword))
    total = q.count()
    posts = (
        q.order_by(TeacherPost.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    items = []
    for p in posts:
        teacher = User.query.get(p.teacher_user_id)
        reqs = CooperationRequest.query.filter_by(post_id=p.id).all()
        selected_students = []
        for r in reqs:
            stu = User.query.get(r.student_user_id)
            selected_students.append(
                {
                    "id": r.id,
                    "student": {
                        "id": stu.id,
                        "display_name": stu.display_name,
                        "username": stu.username,
                    }
                    if stu
                    else None,
                    "teacher_status": r.teacher_status,
                    "student_status": r.student_status,
                    "final_status": r.final_status,
                }
            )
        items.append(
            {
                "id": p.id,
                "post_type": p.post_type,
                "title": p.title,
                "content": p.content,
                "tech_stack": [],
                "tags": [],
                "recruit_count": p.recruit_count,
                "duration": p.duration,
                "outcome": p.outcome,
                "contact": p.contact,
                "deadline": p.deadline.isoformat() if p.deadline else None,
                "teacher": {
                    "id": teacher.id,
                    "display_name": teacher.display_name,
                    "username": teacher.username,
                }
                if teacher
                else None,
                "created_at": p.created_at.isoformat(),
                "selected_students": selected_students,
            }
        )
    return jsonify({"items": items, "total": total, "page": page, "page_size": page_size})


@bp.post("/projects")
@require_roles(["admin"])
def admin_create_project():
    data = request.get_json(force=True)
    title = (data.get("title") or "").strip()
    content = (data.get("content") or "").strip()
    teacher_user_id = data.get("teacher_user_id")
    post_type = (data.get("post_type") or "project").strip()
    if not title or not content or not teacher_user_id:
        return jsonify({"message": "参数不完整"}), 400
    teacher = User.query.get(int(teacher_user_id))
    if not teacher or teacher.role != "teacher":
        return jsonify({"message": "教师不存在"}), 400
    from ..utils import ensure_list_str, json_dumps, now_utc

    post = TeacherPost(
        teacher_user_id=teacher.id,
        post_type=post_type,
        title=title,
        content=content,
        tech_stack_json=json_dumps(ensure_list_str(data.get("tech_stack"))),
        tags_json=json_dumps(ensure_list_str(data.get("tags"))),
        recruit_count=data.get("recruit_count"),
        duration=(data.get("duration") or None),
        outcome=(data.get("outcome") or None),
        contact=(data.get("contact") or None),
        deadline=None,
        visibility="public",
        review_status=ReviewStatus.approved.value,
        created_at=now_utc(),
        updated_at=now_utc(),
    )
    db.session.add(post)
    db.session.commit()
    return jsonify({"id": post.id})


@bp.put("/projects/<int:post_id>")
@require_roles(["admin"])
def admin_update_project(post_id: int):
    data = request.get_json(force=True)
    post = TeacherPost.query.get(post_id)
    if not post:
        return jsonify({"message": "不存在"}), 404
    title = (data.get("title") or post.title or "").strip()
    content = (data.get("content") or post.content or "").strip()
    if not title or not content:
        return jsonify({"message": "标题/内容不能为空"}), 400
    from ..utils import ensure_list_str, json_dumps, now_utc

    post.title = title
    post.content = content
    if "post_type" in data and (data.get("post_type") or "").strip():
        post.post_type = (data.get("post_type") or post.post_type).strip()
    if "tech_stack" in data:
        post.tech_stack_json = json_dumps(ensure_list_str(data.get("tech_stack")))
    if "tags" in data:
        post.tags_json = json_dumps(ensure_list_str(data.get("tags")))
    if "recruit_count" in data:
        post.recruit_count = data.get("recruit_count")
    if "duration" in data:
        post.duration = (data.get("duration") or None)
    if "outcome" in data:
        post.outcome = (data.get("outcome") or None)
    if "contact" in data:
        post.contact = (data.get("contact") or None)
    post.updated_at = now_utc()
    db.session.commit()
    return jsonify({"ok": True})


@bp.delete("/projects/<int:post_id>")
@require_roles(["admin"])
def admin_delete_project(post_id: int):
    post = TeacherPost.query.get(post_id)
    if not post:
        return jsonify({"message": "不存在"}), 404
    db.session.delete(post)
    db.session.commit()
    return jsonify({"ok": True})


@bp.get("/export/users")
@require_roles(["admin"])
def export_users():
    """导出用户数据到Excel"""
    from io import BytesIO
    from flask import send_file
    
    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({"message": "服务器未安装openpyxl库"}), 500
    
    role = (request.args.get("role") or "").strip()
    
    q = User.query
    if role in {"student", "teacher", "admin"}:
        q = q.filter_by(role=role)
    
    users = q.order_by(User.created_at.desc()).all()
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "用户列表"
    
    # 写入表头
    headers = ["角色", "学号/工号", "姓名", "邮箱", "手机号", "注册时间", "状态"]
    ws.append(headers)
    
    # 写入数据
    role_map = {"student": "学生", "teacher": "教师", "admin": "管理员"}
    for u in users:
        ws.append([
            role_map.get(u.role, u.role),
            u.username,
            u.display_name or "",
            u.email or "",
            u.phone or "",
            u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "",
            "启用" if u.is_active else "禁用"
        ])
    
    # 调整列宽
    column_widths = [10, 15, 15, 25, 15, 20, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="用户列表.xlsx"
    )


@bp.get("/export/projects")
@require_roles(["admin"])
def export_projects():
    """导出项目数据到Excel"""
    from io import BytesIO
    from flask import send_file
    
    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({"message": "服务器未安装openpyxl库"}), 500
    
    post_type = (request.args.get("post_type") or "").strip()
    project_status = (request.args.get("project_status") or "").strip()
    
    q = TeacherPost.query
    if post_type in {"project", "innovation", "competition"}:
        q = q.filter_by(post_type=post_type)
    if project_status:
        q = q.filter_by(project_status=project_status)
    
    posts = q.order_by(TeacherPost.created_at.desc()).all()
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "项目列表"
    
    # 写入表头
    headers = ["项目名称", "类型", "教师", "招募人数", "已确认人数", "状态", "创建时间"]
    ws.append(headers)
    
    # 写入数据
    type_map = {"project": "科研项目", "innovation": "大创项目", "competition": "学科竞赛"}
    status_map = {"recruiting": "招募中", "in_progress": "进行中", "completed": "已完成", "closed": "已关闭"}
    
    for p in posts:
        teacher = User.query.get(p.teacher_user_id)
        confirmed_count = CooperationRequest.query.filter_by(
            post_id=p.id,
            final_status=CooperationStatus.confirmed.value
        ).count()
        
        ws.append([
            p.title,
            type_map.get(p.post_type, p.post_type),
            teacher.display_name if teacher else "",
            p.recruit_count or "",
            confirmed_count,
            status_map.get(p.project_status, p.project_status or "招募中"),
            p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else ""
        ])
    
    # 调整列宽
    column_widths = [30, 12, 15, 12, 12, 12, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="项目列表.xlsx"
    )


@bp.post("/import/users")
@require_roles(["admin"])
def import_users():
    """从Excel导入用户数据"""
    from ..utils import hash_password, now_utc
    
    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({"message": "服务器未安装openpyxl库"}), 500
    
    if "file" not in request.files:
        return jsonify({"message": "请上传文件"}), 400
    
    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"message": "请上传Excel文件(.xlsx或.xls)"}), 400
    
    try:
        wb = load_workbook(file)
        ws = wb.active
    except Exception as e:
        return jsonify({"message": f"文件解析失败: {str(e)}"}), 400
    
    # 解析数据
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
    
    results = {
        "total": len(rows),
        "success": 0,
        "failed": 0,
        "errors": []
    }
    
    role_map = {"学生": "student", "教师": "teacher", "管理员": "admin"}
    
    for idx, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        
        try:
            role_text = str(row[0] or "").strip()
            username = str(row[1] or "").strip()
            display_name = str(row[2] or "").strip() if len(row) > 2 else ""
            password = str(row[3] or "123456").strip() if len(row) > 3 else "123456"
            
            # 验证必填字段
            if not role_text:
                results["errors"].append({"row": idx, "message": "角色不能为空"})
                results["failed"] += 1
                continue
            
            if not username:
                results["errors"].append({"row": idx, "message": "学号/工号不能为空"})
                results["failed"] += 1
                continue
            
            # 转换角色
            role = role_map.get(role_text, role_text.lower())
            if role not in {"student", "teacher", "admin"}:
                results["errors"].append({"row": idx, "message": f"无效的角色: {role_text}"})
                results["failed"] += 1
                continue
            
            # 检查用户名是否已存在
            if User.query.filter_by(username=username).first():
                results["errors"].append({"row": idx, "message": f"用户名已存在: {username}"})
                results["failed"] += 1
                continue
            
            # 创建用户
            u = User(
                username=username,
                password_hash=hash_password(password),
                role=role,
                display_name=display_name or username,
                is_active=True,
                created_at=now_utc(),
                must_change_password=password == "123456",
            )
            db.session.add(u)
            results["success"] += 1
            
        except Exception as e:
            results["errors"].append({"row": idx, "message": str(e)})
            results["failed"] += 1
    
    db.session.commit()
    
    return jsonify(results)


@bp.get("/import/users/template")
@require_roles(["admin"])
def get_import_template():
    """获取用户导入模板"""
    from io import BytesIO
    from flask import send_file
    
    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({"message": "服务器未安装openpyxl库"}), 500
    
    wb = Workbook()
    ws = wb.active
    ws.title = "用户导入模板"
    
    # 写入表头
    headers = ["角色", "学号/工号", "姓名", "初始密码(可选)"]
    ws.append(headers)
    
    # 写入示例数据
    ws.append(["学生", "221002501", "张三", "123456"])
    ws.append(["学生", "221002502", "李四", ""])
    ws.append(["教师", "10010001", "王老师", "teacher123"])
    
    # 调整列宽
    column_widths = [10, 15, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="用户导入模板.xlsx"
    )
